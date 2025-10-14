"""
Output Module

This module handles saving data to various formats (CSV, Excel).
Provides unified interface for different output operations.
"""

import logging
import os
import csv
from typing import Dict, List, Any
import openpyxl
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class OutputManager:
    """Manages output operations for different data types and formats."""
    
    def __init__(self, tenant_name: str = None):
        self.default_save_folder = "audit_events_output"
        if not tenant_name:
            raise ValueError("tenant_name is required for OutputManager")
        self.tenant_name = tenant_name
    
    def save_audit_events(self, events: List[Dict], output_format: str, 
                         filename: str = None) -> bool:
        """
        Save audit events to the specified format.
        
        Args:
            events (List[Dict]): List of audit event dictionaries.
            output_format (str): Output format ('csv' or 'excel').
            filename (str, optional): Custom filename. Auto-generated if None.
            
        Returns:
            bool: True if successful, False otherwise.
        """
        if not filename:
            extension = "xlsx" if output_format.lower() == "excel" else output_format.lower()
            filename = f"{self.tenant_name}_audit_events.{extension}"
        
        logger.info(f"Saving {len(events)} audit events to {output_format.upper()} format")
        
        try:
            # Process audit events to extract nested data
            processed_events = self._process_audit_events(events)
            
            if output_format.lower() == "csv":
                return self._write_csv(processed_events, filename)
            elif output_format.lower() == "excel":
                return self._write_excel(processed_events, filename)
            else:
                logger.error(f"Unsupported output format: {output_format}")
                return False
        except Exception as e:
            logger.error(f"Error saving audit events: {e}")
            return False
    
    def save_scan_results(self, scans: List[Dict], output_format: str,
                         filename: str = None, limit: int = None, offset: int = None) -> bool:
        """
        Save scan results to the specified format.
        
        Args:
            scans (List[Dict]): List of scan result dictionaries.
            output_format (str): Output format ('csv' or 'excel').
            filename (str, optional): Custom filename. Auto-generated if None.
            limit (int, optional): Limit value for filename generation.
            offset (int, optional): Offset value for filename generation.
            
        Returns:
            bool: True if successful, False otherwise.
        """
        if not filename:
            # Get proper file extension
            extension = "xlsx" if output_format.lower() == "excel" else output_format.lower()
            if limit is not None and offset is not None:
                filename = f"{self.tenant_name}_scan_results_limit_{limit}_offset_{offset}.{extension}"
            else:
                filename = f"{self.tenant_name}_scan_results.{extension}"
        
        logger.info(f"Saving {len(scans)} scan results to {output_format.upper()} format")
        
        try:
            # Debug: Show first scan raw data
            if scans:
                logger.debug(f"First scan raw data: {scans[0]}")
            
            # Process scans to extract useful fields and handle complex data types
            processed_scans = []
            for i, scan in enumerate(scans):
                processed_scan = self._extract_scan_fields(scan)
                
                if i == 0:  # Debug first scan processing
                    logger.debug(f"Extracted scan fields: {processed_scan}")
                
                # Add remaining fields that weren't extracted
                for key, value in scan.items():
                    if key not in processed_scan:
                        if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
                            processed_scan[key] = "NA"
                            if i == 0:
                                logger.debug(f"Set {key}=NA (was None/empty)")
                        elif isinstance(value, (list, tuple)):
                            processed_scan[key] = ", ".join(str(item) for item in value) if value else "NA"
                            if i == 0:
                                logger.debug(f"Set {key}={processed_scan[key]} (was list/tuple)")
                        elif isinstance(value, dict):
                            processed_scan[key] = str(value) if value else "NA"
                            if i == 0:
                                logger.debug(f"Set {key}={processed_scan[key][:100]}... (was dict)")
                        else:
                            processed_scan[key] = value
                            if i == 0:
                                logger.debug(f"Set {key}={value} (kept as-is)")
                processed_scans.append(processed_scan)
                
                if i == 0:  # Debug first processed scan
                    logger.debug(f"Final processed scan keys: {list(processed_scan.keys())}")
                    logger.debug(f"Sample processed values: {dict(list(processed_scan.items())[:5])}")
            
            if output_format.lower() == "csv":
                return self._write_csv(processed_scans, filename)
            elif output_format.lower() == "excel":
                # Use multi-worksheet Excel format for scans
                return self._save_scans_only_excel(scans, filename)
            else:
                logger.error(f"Unsupported output format: {output_format}")
                return False
        except Exception as e:
            logger.error(f"Error saving scan results: {e}")
            return False
    
    def save_combined_data(self, audit_events: List[Dict], scan_events: List[Dict], 
                          output_format: str, from_date: str = None, to_date: str = None) -> bool:
        """
        Save combined audit events and scan results.
        
        Args:
            audit_events (List[Dict]): List of audit event dictionaries.
            scan_events (List[Dict]): List of scan result dictionaries.
            output_format (str): Output format ('csv' or 'excel').
            from_date (str, optional): Start date for filename.
            to_date (str, optional): End date for filename.
            
        Returns:
            bool: True if successful, False otherwise.
        """
        date_suffix = ""
        if from_date and to_date:
            # Convert dates to safe filename format
            from_safe = from_date.replace("/", "-")
            to_safe = to_date.replace("/", "-")
            date_suffix = f"_{from_safe}_to_{to_safe}"
        
        logger.info(f"Saving combined data: {len(audit_events)} audit events, {len(scan_events)} scan results")
        
        try:
            if output_format.lower() == "csv":
                # Process both data types appropriately
                processed_audit = self._process_audit_events(audit_events) if audit_events else []
                processed_scans = []
                for scan in scan_events:
                    processed_scan = self._extract_scan_fields(scan)
                    # Add remaining fields
                    for key, value in scan.items():
                        if key not in processed_scan:
                            if isinstance(value, (list, tuple)):
                                processed_scan[key] = ", ".join(str(item) for item in value) if value else "NA"
                            elif isinstance(value, dict):
                                processed_scan[key] = str(value) if value else "NA"
                            else:
                                processed_scan[key] = value
                    processed_scans.append(processed_scan)
                
                # Save as separate CSV files
                audit_success = self._write_csv(processed_audit, f"{self.tenant_name}_audit_events{date_suffix}.csv") if processed_audit else True
                scan_success = self._write_csv(processed_scans, f"{self.tenant_name}_scan_results{date_suffix}.csv") if processed_scans else True
                return audit_success and scan_success
                
            elif output_format.lower() == "excel":
                # Save as multi-worksheet Excel file
                return self._save_combined_excel(audit_events, scan_events, f"{self.tenant_name}_cx1_data{date_suffix}.xlsx")
            else:
                logger.error(f"Unsupported output format: {output_format}")
                return False
        except Exception as e:
            logger.error(f"Error saving combined data: {e}")
            return False
    
    def _save_combined_excel(self, audit_events: List[Dict], scan_events: List[Dict], filename: str) -> bool:
        """Save combined data to multi-worksheet Excel file."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import os
        
        try:
            # Ensure output directory exists
            output_folder = "audit_events_output"
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            
            excel_filename = os.path.join(output_folder, filename)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Add audit events worksheet
            if audit_events:
                self._add_audit_worksheet(workbook, audit_events, "Audit Events")
            
            # Group scan events by engine and add worksheets
            if scan_events:
                self._add_scan_worksheets_by_engine(workbook, scan_events)
            
            # Save workbook
            workbook.save(excel_filename)
            print(f"Data successfully written to {excel_filename}")
            return True
            
        except Exception as e:
            print(f"ERROR: Failed to save Excel file: {e}")
            return False
    
    def _add_audit_worksheet(self, workbook, audit_events: List[Dict], sheet_name: str):
        """Add audit events to worksheet."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        worksheet = workbook.create_sheet(title=sheet_name)
        
        # Process audit data same as existing logic
        processed_rows = []
        for row in audit_events:
            new_row = row.copy()
            if "data" in row and isinstance(row["data"], dict):
                data = row["data"]
                new_row["details_id"] = data.get("id", "NA")
                new_row["details_status"] = data.get("status", "NA")
                new_row["details_username"] = data.get("username", "NA")
                new_row.pop("data", None)
            
            # Replace empty values with "NA"
            for key, value in new_row.items():
                if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
                    new_row[key] = "NA"
            processed_rows.append(new_row)
        
        if processed_rows:
            all_columns = set().union(*(row.keys() for row in processed_rows))
            fieldnames = self._order_columns(all_columns)
            
            # Write headers
            for col_num, header in enumerate(fieldnames, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Write data
            for row_num, data_row in enumerate(processed_rows, 2):
                for col_num, header in enumerate(fieldnames, 1):
                    worksheet.cell(row=row_num, column=col_num, value=data_row.get(header, "NA"))
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _add_scan_worksheets_by_engine(self, workbook, scan_events: List[Dict]):
        """Group scans by engine and create separate worksheets."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Group scans by engine
        engines = {}
        for scan in scan_events:
            # Try different possible engine field names
            engine = scan.get("engine", scan.get("scanEngine", scan.get("type", "Unknown Engine")))
            if engine not in engines:
                engines[engine] = []
            engines[engine].append(scan)
        
        # Create worksheet for each engine
        for engine_name, scans in engines.items():
            # Clean sheet name (Excel has restrictions)
            safe_name = str(engine_name).replace("/", "-")[:31]  # Excel sheet name limit
            worksheet = workbook.create_sheet(title=f"Scans")
            
            # Process scan data
            processed_rows = []
            for scan in scans:
                new_row = {}
                
                # First, extract and flatten useful fields
                new_row = self._extract_scan_fields(scan)
                
                # Then handle remaining fields
                for key, value in scan.items():
                    if key not in new_row:  # Don't override extracted fields
                        if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
                            new_row[key] = "NA"
                        elif isinstance(value, (list, tuple)):
                            new_row[key] = ", ".join(str(item) for item in value) if value else "NA"
                        elif isinstance(value, dict):
                            new_row[key] = str(value) if value else "NA"
                        else:
                            new_row[key] = value
                processed_rows.append(new_row)
            
            if processed_rows:
                all_columns = set().union(*(row.keys() for row in processed_rows))
                fieldnames = self._order_columns(all_columns)
                
                # Write headers
                for col_num, header in enumerate(fieldnames, 1):
                    cell = worksheet.cell(row=1, column=col_num, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                
                # Write data
                for row_num, data_row in enumerate(processed_rows, 2):
                    for col_num, header in enumerate(fieldnames, 1):
                        worksheet.cell(row=row_num, column=col_num, value=data_row.get(header, "NA"))
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _extract_scan_fields(self, scan: Dict) -> Dict:
        """Extract and flatten useful fields from scan metadata."""
        import re
        from datetime import datetime
        
        extracted = {}
        logger.debug(f"_extract_scan_fields called with scan keys: {list(scan.keys())}")
        
        # Basic fields
        extracted['scan_id'] = scan.get('id', 'NA')
        
        # Extract repository details from metadata
        metadata = scan.get('metadata', {})
        logger.debug(f"Processing metadata: {metadata}")
        
        if isinstance(metadata, dict):
            # Repository type from metadata
            extracted['repository_type'] = metadata.get('type', 'NA')
            
            # Extract repository details from Handler
            handler = metadata.get('Handler', {})
            if isinstance(handler, dict):
                git_handler = handler.get('GitHandler', {})
                if isinstance(git_handler, dict):
                    # Clean repository URL (remove credentials)
                    repo_url = git_handler.get('repo_url', 'NA')
                    if repo_url != 'NA':
                        # Remove credentials from URL using regex
                        cleaned_url = re.sub(r'https://[^@]*@', 'https://', repo_url)
                        extracted['repository_url'] = cleaned_url
                    else:
                        extracted['repository_url'] = 'NA'
                    
                    # Branch from GitHandler, fallback to scan level
                    extracted['branch'] = git_handler.get('branch', scan.get('branch', 'NA'))
                else:
                    extracted['repository_url'] = 'NA'
                    extracted['branch'] = scan.get('branch', 'NA')
            else:
                extracted['repository_url'] = 'NA'
                extracted['branch'] = scan.get('branch', 'NA')
            
            # Extract project ID from metadata, fallback to scan level
            project = metadata.get('project', {})
            if isinstance(project, dict):
                extracted['project_id'] = project.get('id', scan.get('projectId', 'NA'))
            else:
                extracted['project_id'] = scan.get('projectId', 'NA')
        else:
            # Fallback to scan-level fields if no metadata
            extracted['repository_type'] = 'NA'
            extracted['repository_url'] = 'NA'
            extracted['branch'] = scan.get('branch', 'NA')
            extracted['project_id'] = scan.get('projectId', 'NA')
        
        # Extract scan engines from configs in metadata
        configs = metadata.get('configs', []) if isinstance(metadata, dict) else []
        logger.debug(f"Processing configs: {configs}")
        if isinstance(configs, list):
            enabled_engines = []
            sast_preset = 'NA'
            sca_containers = 'NA'
            microengines_config = 'NA'
            
            for config in configs:
                if isinstance(config, dict):
                    config_type = config.get('type', '')
                    if config_type:
                        enabled_engines.append(config_type)
                    
                    # Extract specific configuration details
                    if config_type == 'sast':
                        value = config.get('value', {})
                        if isinstance(value, dict):
                            preset = value.get('presetName', '')
                            incremental = value.get('incremental', '')
                            sast_preset = f"Preset: {preset}, Incremental: {incremental}" if preset or incremental else 'Default'
                    
                    elif config_type == 'sca':
                        value = config.get('value', {})
                        if isinstance(value, dict):
                            sca_containers = value.get('enableContainersScan', 'false')
                    
                    elif config_type == 'microengines':
                        value = config.get('value', {})
                        if isinstance(value, dict):
                            engines = []
                            for key, val in value.items():
                                if val == 'true':
                                    engines.append(key)
                            microengines_config = ', '.join(engines) if engines else 'None'
            
            extracted['enabled_scan_engines'] = ', '.join(enabled_engines) if enabled_engines else 'NA'
            extracted['sast_configuration'] = sast_preset
            extracted['sca_containers_enabled'] = sca_containers
            extracted['microengines_enabled'] = microengines_config
        else:
            extracted['enabled_scan_engines'] = 'NA'
            extracted['sast_configuration'] = 'NA'
            extracted['sca_containers_enabled'] = 'NA'
            extracted['microengines_enabled'] = 'NA'
        
        # Extract and format created date - prioritize scan-level createdAt
        created_at_str = scan.get('createdAt')
        if created_at_str:
            try:
                # Parse ISO format: 2025-10-10T20:32:35.22344Z
                created_date = datetime.fromisoformat(created_at_str.replace('Z', '+00:00'))
                extracted['created_date'] = created_date.strftime('%m/%d/%Y %H:%M:%S')
            except:
                extracted['created_date'] = 'NA'
        else:
            # Fallback to metadata created_at timestamp
            if isinstance(metadata, dict):
                created_at = metadata.get('created_at', {})
                if isinstance(created_at, dict):
                    seconds = created_at.get('seconds', 0)
                    if seconds:
                        try:
                            created_date = datetime.fromtimestamp(seconds)
                            extracted['created_date'] = created_date.strftime('%m/%d/%Y %H:%M:%S')
                        except:
                            extracted['created_date'] = 'NA'
                    else:
                        extracted['created_date'] = 'NA'
                else:
                    extracted['created_date'] = 'NA'
            else:
                extracted['created_date'] = 'NA'
        
        # Extract engine status details
        status_details = scan.get('statusDetails', [])
        logger.debug(f"statusDetails type: {type(status_details)}, value: {status_details}")
        if isinstance(status_details, list):
            # Initialize all possible engine statuses to NA
            engines = ['general', 'sast', 'sca', 'apisec', 'containers', 'kics', 'microengines']
            for engine in engines:
                extracted[f'{engine}_status'] = 'NA'
                if engine == 'sast':
                    extracted['sast_lines_of_code'] = 'NA'
            logger.debug(f"Initialized {len(engines)} engine statuses to NA")
            
            # Process actual status details
            for status in status_details:
                if isinstance(status, dict):
                    engine_name = status.get('name', '')
                    engine_status = status.get('status', 'NA')
                    
                    if engine_name:
                        # Set engine-specific status
                        extracted[f'{engine_name}_status'] = engine_status
                        
                        # Extract additional metrics
                        if engine_name == 'sast':
                            loc = status.get('loc', 0)
                            extracted['sast_lines_of_code'] = str(loc) if loc else 'NA'
                        
                        # Could add start/end dates if needed:
                        # start_date = status.get('startDate', '')
                        # end_date = status.get('endDate', '')
                        # if start_date:
                        #     extracted[f'{engine_name}_start_date'] = start_date
        else:
            # If no status details, set all to NA
            engines = ['general', 'sast', 'sca', 'apisec', 'containers', 'kics', 'microengines']
            for engine in engines:
                extracted[f'{engine}_status'] = 'NA'
            extracted['sast_lines_of_code'] = 'NA'
        
        logger.debug(f"Final extracted fields count: {len(extracted)}")
        logger.debug(f"Fields with 'NA' values: {[k for k, v in extracted.items() if v == 'NA']}")
        return extracted
    
    def _process_audit_events(self, events: List[Dict]) -> List[Dict]:
        """Process audit events to extract nested data structure."""
        processed_events = []
        
        for event in events:
            new_event = event.copy()
            
            # Handle audit events with nested "data" structure
            if "data" in event:
                data = event["data"]
                if isinstance(data, dict):
                    try:
                        new_event["details_id"] = data.get("id")
                        new_event["details_status"] = data.get("status")
                        new_event["details_username"] = data.get("username")
                    except Exception as e:
                        new_event["details_id"] = None
                        new_event["details_status"] = None
                        new_event["details_username"] = None
                new_event.pop("data", None)  # Remove the nested data field
            
            processed_events.append(new_event)
        
        return processed_events
    
    def _get_logical_column_order(self) -> List[str]:
        """Return logical column order for analyst-friendly Excel output."""
        return [
            # When & What (Critical Context)
            'FormattedEventDate',
            'created_date', 
            'projectName',
            'status',
            
            # Where & How (Source Context)
            'repository_url',
            'branch',
            'repository_type',
            'sourceType',
            'sourceOrigin',
            'initiator',
            
            # What Was Run (Configuration)
            'enabled_scan_engines',
            'sast_configuration',
            'sca_containers_enabled',
            'microengines_enabled',
            
            # How Did It Go (Detailed Results)
            'general_status',
            'sast_status',
            'sast_lines_of_code',
            'sca_status',
            'apisec_status',
            'containers_status',
            'kics_status',
            'microengines_status',
            
            # Technical Details (Deep Dive)
            'scan_id',
            'project_id',
            'createdAt',
            'updatedAt',
            'userAgent',
            'engines',
            'tags',
            'statusDetails',
            'metadata',
            
            # Any other fields will be added alphabetically at the end
        ]
    
    def _order_columns(self, all_columns: set) -> List[str]:
        """Order columns logically for analyst workflow."""
        logical_order = self._get_logical_column_order()
        ordered_columns = []
        
        # Add columns in logical order if they exist
        for col in logical_order:
            if col in all_columns:
                ordered_columns.append(col)
        
        # Add any remaining columns alphabetically at the end
        remaining_columns = sorted(all_columns - set(ordered_columns))
        ordered_columns.extend(remaining_columns)
        
        return ordered_columns
    
    def _save_scans_only_excel(self, scan_events: List[Dict], filename: str) -> bool:
        """Save scan results to multi-worksheet Excel file (scans only)."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import os
        
        try:
            # Ensure output directory exists
            output_folder = "audit_events_output"
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            
            excel_filename = os.path.join(output_folder, filename)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Group scan events by engine and add worksheets
            if scan_events:
                self._add_scan_worksheets_by_engine(workbook, scan_events)
            else:
                # Create empty sheet if no data
                worksheet = workbook.create_sheet(title="No Data")
                worksheet.cell(row=1, column=1, value="No scan results found")
            
            # Save workbook
            workbook.save(excel_filename)
            print(f"Data successfully written to {excel_filename}")
            return True
            
        except Exception as e:
            print(f"ERROR: Failed to save Excel file: {e}")
            return False

    def _write_csv(self, json_list: List[Dict[str, Any]], filename: str) -> bool:
        """Write data to CSV file."""
        if not json_list:
            print("No data to write to CSV.")
            return True

        # Simple processing - replace empty/None values with "NA"
        processed_rows = []
        for row in json_list:
            new_row = row.copy()
            for key, value in new_row.items():
                if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
                    new_row[key] = "NA"
            processed_rows.append(new_row)

        all_columns = set().union(*(row.keys() for row in processed_rows))
        fieldnames = self._order_columns(all_columns)
        
        # Ensure the target folder exists
        if not os.path.exists(self.default_save_folder):
            os.makedirs(self.default_save_folder)

        csv_filename = os.path.join(self.default_save_folder, filename)

        # Write to CSV with error handling
        try:
            with open(csv_filename, 'w', newline='', encoding='utf-8') as output_file:
                dict_writer = csv.DictWriter(output_file, fieldnames=fieldnames)
                dict_writer.writeheader()
                dict_writer.writerows(processed_rows)
            print(f"Data successfully written to {csv_filename}")
            return True
        except PermissionError:
            print(f"ERROR: Cannot save {csv_filename}")
            print(f"   - File may be open in another program")
            print(f"   - Close the file and try again")
            return False
        except Exception as e:
            print(f"ERROR: Failed to save CSV file: {e}")
            print(f"   - Check file permissions")
            return False

    def _write_excel(self, json_list: List[Dict[str, Any]], filename: str) -> bool:
        """Write data to Excel file."""
        if not json_list:
            print("No data to write to Excel.")
            return True

        # Simple processing - replace empty/None values with "NA"
        processed_rows = []
        for row in json_list:
            new_row = row.copy()
            for key, value in new_row.items():
                if value is None or value == "" or (isinstance(value, str) and value.strip() == ""):
                    new_row[key] = "NA"
            processed_rows.append(new_row)

        all_columns = set().union(*(row.keys() for row in processed_rows))
        fieldnames = self._order_columns(all_columns)

        # Ensure the target folder exists
        if not os.path.exists(self.default_save_folder):
            os.makedirs(self.default_save_folder)

        excel_filename = os.path.join(self.default_save_folder, filename)

        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Audit Trail Results"

        # Write headers
        for col_num, header in enumerate(fieldnames, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Write data rows
        for row_num, data_row in enumerate(processed_rows, 2):
            for col_num, header in enumerate(fieldnames, 1):
                worksheet.cell(row=row_num, column=col_num, value=data_row.get(header, "NA"))

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save the workbook with error handling
        try:
            workbook.save(excel_filename)
            print(f"Data successfully written to {excel_filename}")
            return True
        except PermissionError:
            print(f"ERROR: Cannot save {excel_filename}")
            print(f"   - File may be open in Excel or another program")
            print(f"   - Close the file and try again")
            print(f"   - Or use --output csv as an alternative")
            return False
        except FileNotFoundError:
            print(f"ERROR: Directory path not found for {excel_filename}")
            print(f"   - Check if the directory exists")
            return False
        except Exception as e:
            print(f"ERROR: Failed to save Excel file: {e}")
            print(f"   - Try using --output csv instead")
            print(f"   - Or check file permissions")
            return False

    def get_supported_formats(self) -> List[str]:
        """Get list of supported output formats."""
        return ["csv", "excel"]
